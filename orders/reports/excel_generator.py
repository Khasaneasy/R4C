import openpyxl
from openpyxl.utils import get_column_letter
from datetime import timedelta, date

from django.db.models import Count

from robots.models import Robot


def excel_generate():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Сводка'

    headers = ['Модель', 'Версия', 'Количество']
    for col_num, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_num)
        sheet[f'{column_letter}1'] = header

    last_week = date.today() - timedelta(days=7)
    data = (
        Robot.objects.filter(created__gte=last_week)
        .values('model', 'version')
        .annotate(count=Count('id'))
    )

    for row_num, entry in enumerate(data, start=2):
        sheet[f"A{row_num}"] = entry["robot__model"]
        sheet[f"B{row_num}"] = entry["robot__version"]
        sheet[f"C{row_num}"] = entry["count"]

    return wb
